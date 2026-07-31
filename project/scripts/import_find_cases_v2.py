from __future__ import annotations

import json
import re
import warnings
from pathlib import Path
from typing import Any

import openpyxl
import yaml


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

ROOT = Path(__file__).resolve().parents[1]
DESKTOP = Path("C:/Users/Administrator/Desktop")
CASE_DATA_PATH = ROOT / "testdata" / "find_task_cases.yaml"
BASE_METHODS_PATH = ROOT / "testdata" / "base_methods.yaml"


def main() -> None:
    excel_path = latest_excel()
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    cases = build_cases(workbook["testcases"])
    base_methods = build_base_methods(workbook["bases"], excel_path)

    CASE_DATA_PATH.write_text(
        yaml.safe_dump(
            {
                "source_excel": str(excel_path),
                "general_rules": read_general_rules(workbook["general"]),
                "cases": cases,
            },
            allow_unicode=True,
            sort_keys=False,
            width=140,
        ),
        encoding="utf-8",
    )
    BASE_METHODS_PATH.write_text(
        yaml.safe_dump(base_methods, allow_unicode=True, sort_keys=False, width=140),
        encoding="utf-8",
    )
    print(f"Imported {len(cases)} cases from {excel_path}")


def latest_excel() -> Path:
    files = sorted(DESKTOP.glob("Find*2.0.xlsx"), key=lambda path: path.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError("No Find*2.0.xlsx file found on Desktop")
    return files[0]


def read_general_rules(sheet) -> list[str]:
    return [str(row[1]) for row in sheet.iter_rows(min_row=2, values_only=True) if row[1]]


def build_base_methods(sheet, excel_path: Path) -> dict[str, Any]:
    methods = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        methods.append(
            {
                "name": row[0],
                "steps": row[1],
                "request_example": row[2],
                "response_example": row[3],
                "purpose": row[4],
            }
        )
    return {"source_excel": str(excel_path), "methods": methods}


def build_cases(sheet) -> list[dict[str, Any]]:
    cases = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        case_id, title, steps, parse_example, payload_text, assertion_text = row[:6]
        endpoint = infer_endpoint(payload_text or "", steps or "")
        payload = parse_payload(payload_text or "")
        if endpoint == "/api/smartPluginTask/build":
            payload = normalize_payload(payload)
        else:
            payload = normalize_placeholders(payload)
        case = {
            "id": case_id,
            "title": title or "",
            "platform": infer_platform(case_id, payload),
            "steps": steps or "",
            "parse_target_example": parse_example or "",
            "request": {
                "method": "POST",
                "endpoint": endpoint,
                "json": payload,
            },
            "assertions": build_assertions(case_id, title or "", assertion_text or "", endpoint),
        }
        prepare_rules = infer_prepare_rules(case_id, title or "", steps or "", payload_text or "", assertion_text or "")
        if prepare_rules:
            case["prepare_variables"] = prepare_rules
            case["depends_on"] = sorted({rule.get("source_case") for rule in prepare_rules if rule.get("source_case")})
        cases.append(case)
    return cases


def infer_endpoint(payload_text: str, steps: str) -> str:
    text = f"{payload_text}\n{steps}"
    match = re.search(r"(/api/[A-Za-z0-9_/]+)", text)
    return match.group(1) if match else "/api/smartPluginTask/build"


def parse_payload(text: str) -> dict[str, Any]:
    start = text.find("{")
    if start < 0:
        raise ValueError(f"No JSON object found in payload text: {text[:200]}")
    raw = text[start:]
    raw = raw.replace('${fdname}",', '${fdname}\\",')
    raw = raw.replace("${avg1]", "${avg1}")
    raw = re.sub(r"\$\{avg1][^}]+}", "${avg1}", raw)
    raw = re.sub(r"\$\{\s*\}", '"__PLACEHOLDER__"', raw)
    raw = re.sub(r"\$\{\s*([A-Za-z][A-Za-z0-9_]*)\s*\}", r"${\1}", raw)
    return json.loads(raw)


def normalize_payload(payload: dict[str, Any]) -> dict[str, Any]:
    payload = normalize_placeholders(payload)
    payload["fdOriginalThirdId"] = "${thirdid}"
    payload["fdOriginalDeviceId"] = "${deviceid}"
    payload["fdName"] = "${fdName}"
    return payload


def normalize_placeholders(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: normalize_placeholders(item) for key, item in value.items()}
    if isinstance(value, list):
        return [normalize_placeholders(item) for item in value]
    if value == "__PLACEHOLDER__":
        return "${runtime}"
    if isinstance(value, str):
        return re.sub(r"\$\{\s*([A-Za-z][A-Za-z0-9_]*)\s*\}", r"${\1}", value)
    return value


def infer_platform(case_id: str, payload: dict[str, Any]) -> str:
    if case_id.startswith("xhs"):
        return "xhs"
    if case_id.startswith("dy"):
        return "dy"
    if case_id.startswith("xt"):
        return "xt"
    if case_id.startswith("pgy"):
        return "pgy"
    return "xt" if payload.get("fdType") in (1, 3, "1", "3") else "pgy"


def build_assertions(case_id: str, title: str, assertion_text: str, endpoint: str = "/api/smartPluginTask/build") -> dict[str, Any]:
    assertions: dict[str, Any] = {"build_success": True, "item_checks": []}
    if endpoint == "/api/collectTask/buildCollectTask":
        assertions["collect_status_success"] = True
        detail_checks = parse_collect_detail_assertions(assertion_text)
        if detail_checks:
            assertions["collect_detail_required"] = True
            assertions["item_checks"].extend(detail_checks)
        return assertions
    result_var, id_field = parse_result_variable(assertion_text)
    if result_var and id_field:
        assertions["item_count_equals_variable_length"] = result_var
        assertions["item_checks"].append({"field": id_field, "in_variable": result_var})
    else:
        assertions["item_checks"].extend(parse_contains_assertions(assertion_text))
    return assertions


def parse_collect_detail_assertions(text: str) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    in_match = re.search(r"(fd[A-Za-z0-9_]+)\s+in\s*\(([^)]+)\)", text)
    if in_match:
        values = [part.strip().strip("'\"") for part in in_match.group(2).split(",") if part.strip()]
        checks.append({"field": in_match.group(1), "in_values": values})
    for field, value in re.findall(r"(fd[A-Za-z0-9_]+)\s*=\s*[\"']([^\"']+)[\"']", text):
        checks.append({"field": field, "equals": value})
    return checks


def parse_contains_assertions(text: str) -> list[dict[str, Any]]:
    checks = []
    for field in ("fdContentTags", "fdLabels"):
        if field not in text:
            continue
        field_index = text.find(field)
        tail = text[field_index : field_index + 300]
        keyword_match = re.search(r"包含[“\"]([^”\"]+)[”\"]关键字", tail)
        if keyword_match:
            checks.append({"field": field, "contains": [keyword_match.group(1)]})
    return checks


def parse_result_variable(text: str) -> tuple[str | None, str | None]:
    match = re.search(r"达人数量\s*=\s*([A-Za-z][A-Za-z0-9_]*)列表length.*?(fdcode|fdCode|fdUid|fduid)", text, re.I | re.S)
    if not match:
        return None, None
    field = "fdUid" if "uid" in match.group(2).lower() else "fdCode"
    return match.group(1), field


def infer_prepare_rules(case_id: str, title: str, steps: str, payload_text: str, assertion_text: str) -> list[dict[str, Any]]:
    placeholders = {name.strip() for name in re.findall(r"\$\{\s*([A-Za-z][A-Za-z0-9_]*)\s*\}", payload_text)}
    result_var, id_field = parse_result_variable(assertion_text)
    if not result_var:
        return []

    source_case = first_source_case(steps, case_id)
    rules: list[dict[str, Any]] = []
    id_field = id_field or ("fdUid" if case_id.startswith("xt") else "fdCode")

    if {"tage1", "tage2"} & placeholders:
        rules.append(value_filter_rule(source_case, "fdFeatureTags", id_field, "split", placeholders, result_var, title, steps))
    elif {"label1", "label2"} & placeholders:
        rules.append(value_filter_rule(source_case, "fdLabels", id_field, "labels", placeholders, result_var, title, steps))
    elif {"trade1", "trade2"} & placeholders:
        rules.append(value_filter_rule(source_case, "fdTopTradeType", id_field, "split", placeholders, result_var, title, steps))
    elif "province1" in placeholders:
        rules.append(distribution_rule(source_case, "fdProvincesDistributionOfFanPortrait", "province1", result_var, id_field, 1))
    elif "city1" in placeholders:
        rules.append(distribution_rule(source_case, "fdCitiesDistributionOfFanPortrait", "city1", result_var, id_field, 1))
    elif "device1" in placeholders:
        rules.append(distribution_rule(source_case, "fdDevicesDistributionOfFanPortrait", "device1", result_var, id_field, 2))
    elif "cmhotkey1" in placeholders:
        rules.append(value_filter_rule(source_case, "fdCommentHotKey", id_field, "hotkey", placeholders, result_var, title, steps))
    elif "cohotkey1" in placeholders:
        rules.append(value_filter_rule(source_case, "fdContentHotKey", id_field, "hotkey", placeholders, result_var, title, steps))
    elif {"fdname", "id"} & placeholders:
        rules.append(task_difference_rule(case_id, steps, result_var, id_field))
    elif "fdMcnName" in steps:
        rules.append(empty_filter_rule(source_case, "fdMcnName", id_field, result_var, should_be_empty="非机构" in title))
    elif numeric_field := first_numeric_field(steps):
        rules.append(numeric_rule(source_case, numeric_field, id_field, result_var, title, placeholders))
    return rules


def first_source_case(steps: str, current_case: str) -> str:
    prefix = "xt" if current_case.startswith("xt") else "pgy"
    matches = re.findall(rf"{prefix}\d+", steps)
    return matches[0] if matches else ("xt05" if prefix == "xt" else "pgy05")


def value_filter_rule(source_case: str, field: str, id_field: str, extractor: str, placeholders: set[str], result_var: str, title: str, steps: str) -> dict[str, Any]:
    value_vars = []
    include_vars = []
    exclude_vars = []
    if any(name.endswith("1") for name in placeholders):
        first = sorted(name for name in placeholders if name.endswith("1"))[0]
        value_vars.append({"name": first, "rank": 1, "order": "desc"})
        include_vars.append(first)
    if any(name.endswith("2") for name in placeholders):
        second = sorted(name for name in placeholders if name.endswith("2"))[0]
        second_rule = {"name": second, "rank": 1, "order": "desc"}
        if "最少" in steps:
            second_rule.update({"rank": 1, "order": "asc"})
        elif "第二" in steps:
            second_rule.update({"rank": 2, "order": "desc"})
        if field == "fdFeatureTags" and "除tage1" in steps:
            second_rule["scope_include_variables"] = include_vars[:]
            second_rule["exclude_variables"] = include_vars[:]
        value_vars.append(second_rule)
        exclude_vars.append(second)
    if "不包含" in title and len(value_vars) == 1:
        include_vars = []
        exclude_vars = [value_vars[0]["name"]]
    return {
        "type": "value_filter",
        "source_case": source_case,
        "source_field": field,
        "id_field": id_field,
        "extractor": extractor,
        "value_variables": value_vars,
        "include_variables": include_vars,
        "exclude_variables": exclude_vars,
        "result_variable": result_var,
    }


def distribution_rule(source_case: str, field: str, value_var: str, result_var: str, id_field: str, rank: int) -> dict[str, Any]:
    return {
        "type": "value_filter",
        "source_case": source_case,
        "source_field": field,
        "id_field": id_field,
        "extractor": "distribution_top3",
        "value_variables": [{"name": value_var, "rank": rank}],
        "include_variables": [value_var],
        "exclude_variables": [],
        "result_variable": result_var,
    }


def empty_filter_rule(source_case: str, field: str, id_field: str, result_var: str, should_be_empty: bool) -> dict[str, Any]:
    return {
        "type": "empty_filter",
        "source_case": source_case,
        "source_field": field,
        "id_field": id_field,
        "should_be_empty": should_be_empty,
        "result_variable": result_var,
    }


def first_numeric_field(steps: str) -> str | None:
    for field in (
        "fdPgyThirtyCooperateNoteNumber",
        "fdPgyNinetyCooperateNoteNumber",
        "fdPgyNinetyNoteNumber",
        "fdPgyNoteNumber",
    ):
        if field in steps:
            return field
    return None


def numeric_rule(source_case: str, field: str, id_field: str, result_var: str, title: str, placeholders: set[str]) -> dict[str, Any]:
    if "区间" in title:
        relation = "between_avg_max"
    elif "小于" in title:
        relation = "lte_avg"
    else:
        relation = "gte_avg"
    return {
        "type": "numeric_filter",
        "source_case": source_case,
        "source_field": field,
        "id_field": id_field,
        "result_variable": result_var,
        "relation": relation,
        "average_variable": "avg1" if "avg1" in placeholders else None,
        "max_variable": "max1" if "max1" in placeholders else None,
    }


def task_difference_rule(case_id: str, steps: str, result_var: str, id_field: str) -> dict[str, Any]:
    prefix = "xt" if case_id.startswith("xt") else "pgy"
    cases = re.findall(rf"{prefix}\d+", steps)
    meta_case = cases[0] if cases else None
    include_case = cases[1] if len(cases) > 1 else ("xt05" if prefix == "xt" else "pgy01")
    exclude_case = cases[0] if cases else None
    return {
        "type": "task_difference",
        "source_case": include_case,
        "exclude_case": exclude_case,
        "metadata_case": meta_case,
        "id_field": id_field,
        "result_variable": result_var,
        "fdname_variable": "fdname",
        "id_variable": "id",
    }


if __name__ == "__main__":
    main()
